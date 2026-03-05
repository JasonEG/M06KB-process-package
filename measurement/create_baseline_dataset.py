"""
Create measurement/baseline-dataset.xlsx
M08 — Grade Finalization Measurement & Baseline Dataset
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from datetime import date
import os

# ---------------------------------------------------------------------------
# Raw data (mirrors Q1-finalization-export.csv)
# ---------------------------------------------------------------------------
RAW_HEADERS = [
    "TeacherID", "Department", "NumClasses", "FinalizationDate",
    "Deadline", "WeightErrors", "NotificationComplete",
    "ReworkIncidents", "AvgTimePerClass_min"
]

RAW_ROWS = [
    ("T001","Math",   4, date(2025,10,29), date(2025,10,31), 0, 1, 0,  7.5),
    ("T002","Math",   5, date(2025,10,30), date(2025,10,31), 0, 1, 0,  6.8),
    ("T003","Math",   3, date(2025,11, 2), date(2025,10,31), 1, 0, 1, 11.2),
    ("T004","Math",   4, date(2025,10,28), date(2025,10,31), 0, 1, 0,  8.1),
    ("T005","Science",5, date(2025,10,31), date(2025,10,31), 1, 1, 1,  9.4),
    ("T006","Science",4, date(2025,10,30), date(2025,10,31), 0, 0, 0,  7.2),
    ("T007","Science",3, date(2025,11, 3), date(2025,10,31), 0, 1, 2, 12.5),
    ("T008","Science",4, date(2025,10,29), date(2025,10,31), 1, 1, 0,  8.0),
    ("T009","English",5, date(2025,10,31), date(2025,10,31), 0, 1, 0,  6.5),
    ("T010","English",4, date(2025,10,28), date(2025,10,31), 0, 0, 1,  7.8),
    ("T011","English",3, date(2025,10,30), date(2025,10,31), 1, 1, 0,  9.1),
    ("T012","English",4, date(2025,11, 1), date(2025,10,31), 0, 0, 0, 10.3),
    ("T013","History",3, date(2025,10,29), date(2025,10,31), 0, 1, 0,  7.0),
    ("T014","History",4, date(2025,10,30), date(2025,10,31), 0, 1, 1,  8.6),
    ("T015","History",5, date(2025,10,31), date(2025,10,31), 0, 0, 0,  7.3),
    ("T016","Electives",2,date(2025,10,27),date(2025,10,31),0, 1, 0,  5.5),
    ("T017","Electives",3,date(2025,10,29),date(2025,10,31),1, 0, 1, 11.8),
    ("T018","Electives",2,date(2025,10,30),date(2025,10,31),0, 1, 0,  6.2),
]

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
HEADER_FILL  = PatternFill("solid", fgColor="2F4F8F")   # dark blue
ALT_FILL     = PatternFill("solid", fgColor="EEF2FA")   # light blue stripe
BORDER_THIN  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
GREEN_FONT  = Font(color="276221", bold=True, name="Calibri", size=11)
YELLOW_FONT = Font(color="9C6500", bold=True, name="Calibri", size=11)
RED_FONT    = Font(color="9C0006", bold=True, name="Calibri", size=11)

TITLE_FONT  = Font(bold=True, name="Calibri", size=13)
LABEL_FONT  = Font(bold=True, name="Calibri", size=11)
BODY_FONT   = Font(name="Calibri", size=11)

def style_header_row(ws, row, col_start, col_end):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN

def style_data_cell(ws, row, col, alt=False):
    cell = ws.cell(row=row, column=col)
    cell.font   = BODY_FONT
    cell.border = BORDER_THIN
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if alt:
        cell.fill = ALT_FILL

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

# ---------------------------------------------------------------------------
# Sheet 1 — Raw Data
# ---------------------------------------------------------------------------
def build_raw_sheet(wb):
    ws = wb.active
    ws.title = "Raw Data"

    ws.row_dimensions[1].height = 18
    ws["A1"] = "Q1 Grade Finalization Export — Raw Data (Power Query source, do not edit)"
    ws["A1"].font = TITLE_FONT

    # Headers at row 2
    for c, h in enumerate(RAW_HEADERS, 1):
        ws.cell(row=2, column=c, value=h)
    style_header_row(ws, 2, 1, len(RAW_HEADERS))

    # Data
    for r, row in enumerate(RAW_ROWS, 3):
        alt = (r % 2 == 0)
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
            style_data_cell(ws, r, c, alt)
            # Format dates
            if c in (4, 5):
                ws.cell(row=r, column=c).number_format = "MM/DD/YYYY"

    # Column widths
    widths = [10, 11, 12, 18, 12, 14, 22, 18, 22]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, i, w)

    ws.freeze_panes = "A3"

# ---------------------------------------------------------------------------
# Sheet 2 — Cleaned Data
# ---------------------------------------------------------------------------
def build_cleaned_sheet(wb):
    ws = wb.create_sheet("Cleaned Data")

    CLEAN_HEADERS = RAW_HEADERS + ["DaysFromDeadline", "OnTime", "WeightErrorFlag"]

    ws.row_dimensions[1].height = 18
    ws["A1"] = "Cleaned Data — Power Query transformations applied; calculated columns added"
    ws["A1"].font = TITLE_FONT

    for c, h in enumerate(CLEAN_HEADERS, 1):
        ws.cell(row=2, column=c, value=h)
    style_header_row(ws, 2, 1, len(CLEAN_HEADERS))

    deadline = date(2025, 10, 31)
    for r, row in enumerate(RAW_ROWS, 3):
        alt = (r % 2 == 0)
        fin_date = row[3]
        days_from = (fin_date - deadline).days
        on_time = 1 if days_from <= 0 else 0
        weight_flag = 1 if row[5] > 0 else 0
        full_row = row + (days_from, on_time, weight_flag)

        for c, val in enumerate(full_row, 1):
            ws.cell(row=r, column=c, value=val)
            style_data_cell(ws, r, c, alt)
            if c in (4, 5):
                ws.cell(row=r, column=c).number_format = "MM/DD/YYYY"

        # Highlight late rows (DaysFromDeadline > 0) in column 10 with red tint
        if days_from > 0:
            ws.cell(row=r, column=10).fill = PatternFill("solid", fgColor="FFC7CE")

    widths = [10, 11, 12, 18, 12, 14, 22, 18, 22, 18, 9, 16]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, i, w)

    ws.freeze_panes = "A3"

    # Notes block below data
    note_row = len(RAW_ROWS) + 4
    ws.cell(row=note_row, column=1, value="Column Notes:")
    ws.cell(row=note_row, column=1).font = LABEL_FONT
    notes = [
        ("DaysFromDeadline", "FinalizationDate − Deadline  |  Negative = early, 0 = on time, Positive = late"),
        ("OnTime",           "1 if DaysFromDeadline ≤ 0, else 0"),
        ("WeightErrorFlag",  "1 if WeightErrors > 0, else 0"),
    ]
    for i, (col_name, note) in enumerate(notes, 1):
        ws.cell(row=note_row + i, column=1, value=col_name).font = Font(bold=True, name="Calibri", size=10)
        ws.cell(row=note_row + i, column=2, value=note).font = Font(name="Calibri", size=10)

# ---------------------------------------------------------------------------
# Sheet 3 — Summary Pivot
# ---------------------------------------------------------------------------
def build_pivot_sheet(wb):
    ws = wb.create_sheet("Summary Pivot")

    # Pre-computed department summaries
    dept_data = {
        "Math":     {"teachers": ["T001","T002","T003","T004"], "avg_time": 8.40, "weight_err": 1, "notif_pct": 75.0, "ontime_pct": 75.0},
        "Science":  {"teachers": ["T005","T006","T007","T008"], "avg_time": 9.28, "weight_err": 2, "notif_pct": 75.0, "ontime_pct": 75.0},
        "English":  {"teachers": ["T009","T010","T011","T012"], "avg_time": 8.43, "weight_err": 1, "notif_pct": 50.0, "ontime_pct": 75.0},
        "History":  {"teachers": ["T013","T014","T015"],        "avg_time": 7.63, "weight_err": 0, "notif_pct": 66.7, "ontime_pct":100.0},
        "Electives":{"teachers": ["T016","T017","T018"],        "avg_time": 7.83, "weight_err": 1, "notif_pct": 66.7, "ontime_pct":100.0},
    }

    ws.row_dimensions[1].height = 18
    ws["A1"] = "Summary by Department — PivotTable (simulated; Power Query source)"
    ws["A1"].font = TITLE_FONT

    pivot_headers = [
        "Department", "Teachers (n)", "Avg Time/Class (min)",
        "Total Weight Errors", "Notification Complete (%)", "On-Time Rate (%)"
    ]
    for c, h in enumerate(pivot_headers, 1):
        ws.cell(row=2, column=c, value=h)
    style_header_row(ws, 2, 1, len(pivot_headers))

    departments = ["Math", "Science", "English", "History", "Electives"]
    for r, dept in enumerate(departments, 3):
        d = dept_data[dept]
        alt = (r % 2 == 0)
        row_vals = [dept, len(d["teachers"]), d["avg_time"], d["weight_err"], d["notif_pct"], d["ontime_pct"]]
        for c, val in enumerate(row_vals, 1):
            ws.cell(row=r, column=c, value=val)
            style_data_cell(ws, r, c, alt)
            if c in (3,):
                ws.cell(row=r, column=c).number_format = "0.00"
            if c in (5, 6):
                ws.cell(row=r, column=c).number_format = "0.0"

    # Grand total row
    total_row = 8
    ws.cell(row=total_row, column=1, value="ALL DEPARTMENTS").font = Font(bold=True, name="Calibri", size=11)
    ws.cell(row=total_row, column=2, value=18)
    ws.cell(row=total_row, column=3, value=8.40).number_format = "0.00"
    ws.cell(row=total_row, column=4, value=5)
    ws.cell(row=total_row, column=5, value=66.7).number_format = "0.0"
    ws.cell(row=total_row, column=6, value=83.3).number_format = "0.0"
    for c in range(1, 7):
        ws.cell(row=total_row, column=c).border = BORDER_THIN
        ws.cell(row=total_row, column=c).fill = PatternFill("solid", fgColor="D9E1F2")
        ws.cell(row=total_row, column=c).alignment = Alignment(horizontal="center")
        ws.cell(row=total_row, column=c).font = Font(bold=True, name="Calibri", size=11)

    widths = [14, 14, 22, 22, 24, 18]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, i, w)

    # --- Bar Chart: Avg Time-to-Finalize by Department ---
    chart = BarChart()
    chart.type = "col"
    chart.title = "Avg Time-to-Finalize per Class by Department (min)"
    chart.y_axis.title = "Minutes"
    chart.x_axis.title = "Department"
    chart.style = 10
    chart.height = 12
    chart.width = 18

    data_ref = Reference(ws, min_col=3, min_row=2, max_row=7)
    cats_ref = Reference(ws, min_col=1, min_row=3, max_row=7)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # Add target line annotation (as a second series — just a reference)
    ws.cell(row=10, column=1, value="Target (≤ 8.0 min)").font = Font(italic=True, name="Calibri", size=10, color="9C0006")

    ws.add_chart(chart, "A11")

# ---------------------------------------------------------------------------
# Sheet 4 — Dashboard
# ---------------------------------------------------------------------------
def build_dashboard_sheet(wb):
    ws = wb.create_sheet("Dashboard")

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 18

    # Title block
    ws.merge_cells("B1:F1")
    ws["B1"] = "Grade Finalization — Q1 Baseline Dashboard"
    ws["B1"].font = Font(bold=True, name="Calibri", size=16, color="2F4F8F")
    ws["B1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("B2:F2")
    ws["B2"] = "Period: Q1 Fall 2025  |  Deadline: 10/31/2025  |  n = 18 teachers  |  Source: Q1-finalization-export.csv"
    ws["B2"].font = Font(italic=True, name="Calibri", size=10, color="595959")
    ws.row_dimensions[2].height = 16

    # KPI table header
    kpi_hdr_row = 4
    hdr_vals = ["Metric ID", "Metric Name", "Q1 Baseline", "Improvement Target", "Gap", "Status"]
    for c, h in enumerate(hdr_vals, 2):
        cell = ws.cell(row=kpi_hdr_row, column=c, value=h)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER_THIN
    ws.row_dimensions[kpi_hdr_row].height = 20

    # KPI rows
    kpi_data = [
        ("M-01", "Finalization completion rate",    "83.3%",             "≥ 95%",        "−11.7 pp",           "RED"),
        ("M-02", "Weight configuration error rate", "0.28 errors/tchr",  "≤ 0.10",       "+0.18 errors/tchr",  "RED"),
        ("M-03", "Notification compliance rate",    "66.7%",             "≥ 90%",        "−23.3 pp",           "RED"),
        ("M-04", "Avg time-to-finalize per class",  "8.4 min",           "≤ 8.0 min",    "+0.4 min",           "YELLOW"),
        ("M-05", "Rework incident rate",            "0.39 incidents/tchr","≤ 0.20",      "+0.19 incidents/tchr","RED"),
    ]

    status_style = {
        "GREEN":  (GREEN_FILL,  GREEN_FONT,  "MEETS TARGET"),
        "YELLOW": (YELLOW_FILL, YELLOW_FONT, "NEAR TARGET"),
        "RED":    (RED_FILL,    RED_FONT,    "MISSES TARGET"),
    }

    for i, (mid, name, baseline, target, gap, status) in enumerate(kpi_data):
        r = kpi_hdr_row + 1 + i
        ws.row_dimensions[r].height = 22
        alt = (i % 2 == 1)
        base_fill = ALT_FILL if alt else PatternFill("solid", fgColor="FFFFFF")

        vals = [mid, name, baseline, target, gap]
        for c, val in enumerate(vals, 2):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font      = BODY_FONT
            cell.fill      = base_fill
            cell.border    = BORDER_THIN
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Status cell (col G = 7)
        sfill, sfont, slabel = status_style[status]
        scell = ws.cell(row=r, column=7, value=slabel)
        scell.font      = sfont
        scell.fill      = sfill
        scell.border    = BORDER_THIN
        scell.alignment = Alignment(horizontal="center", vertical="center")

    # Legend
    legend_row = kpi_hdr_row + len(kpi_data) + 3
    ws.cell(row=legend_row, column=2, value="Legend:").font = LABEL_FONT
    legend_items = [
        (GREEN_FILL,  GREEN_FONT,  "MEETS TARGET    At or better than target"),
        (YELLOW_FILL, YELLOW_FONT, "NEAR TARGET     Within 10% of target value"),
        (RED_FILL,    RED_FONT,    "MISSES TARGET   More than 10% from target"),
    ]
    for j, (lfill, lfont, ltext) in enumerate(legend_items, 1):
        cell = ws.cell(row=legend_row + j, column=2, value=ltext)
        cell.fill   = lfill
        cell.font   = lfont
        cell.border = BORDER_THIN
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[legend_row + j].height = 18

    # OP code mapping note
    note_row = legend_row + len(legend_items) + 2
    ws.cell(row=note_row, column=2, value="OP Code Mapping:").font = LABEL_FONT
    op_notes = [
        "M-01 → OP-01 (manual class cycling), OP-09 (ambiguous start trigger)",
        "M-02 → OP-02 (no weight validation reference)",
        "M-03 → OP-03 (unstructured email notification)",
        "M-04 → OP-01 (manual class cycling without completion tracker)",
        "M-05 → OP-04 (save-before-exit omission), OP-05 (no acceptance criteria for Step 7)",
    ]
    for k, note in enumerate(op_notes, 1):
        ws.cell(row=note_row + k, column=2, value=note).font = Font(name="Calibri", size=10, italic=True)

    # Adjust status column width
    ws.column_dimensions["G"].width = 18

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    out_path = os.path.join(os.path.dirname(__file__), "baseline-dataset.xlsx")

    wb = openpyxl.Workbook()
    build_raw_sheet(wb)
    build_cleaned_sheet(wb)
    build_pivot_sheet(wb)
    build_dashboard_sheet(wb)

    wb.save(out_path)
    print(f"Saved: {out_path}")
